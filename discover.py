#!/usr/bin/env python
__version__ = "0.0.1"
__author__ = "Utkarsh Mahar"
__author_email__ = "umahar@cisco.com"
__copyright__ = "Copyright (c) 2021 Cisco Systems. All rights reserved."
__license__ = "MIT"

import csv
import os
import pprint,requests
import sys
import yaml
from datetime import datetime
from genie.conf.base import Device
from genie.testbed import load
from mac_vendor_lookup import MacLookup
from openpyxl import Workbook, load_workbook
from webexteamssdk import WebexTeamsAPI
import argparse
date_time_now =  datetime.now().strftime("%d_%m_%H_%M")
import openpyxl
import ipaddress
import time
import zipfile
parser = argparse.ArgumentParser()
parser.add_argument('--site', type=str, required=True, help="Site name")
parser.add_argument('--file', type=str, required=True, help="Input filename")
parser.add_argument('--webex', type=str, required=False, help="Webex file for publishing")
parser.add_argument('--ssh_options', type=str, required=False, help="SSH Options for old ciphers")
parser.add_argument('--offline', action="store_true", required=False, help="Offline sites")

args = parser.parse_args()
site = args.site
file = args.file
webex = args.webex
ssh_options = args.ssh_options
offline_flag = args.offline

class Discover():

    def __init__(self,site,file):
        self.site = site
        self.switches = []
        self.device_connection_objs = {}
        self.connected_devices = []
        self.not_connected_devices = []
        if args.webex:
           self.download_file()        
        self.process_input_file(file)
        if offline_flag:
           self.build_offline_testbed()
        else:
         self.build_testbed()
        self.testbed = load('testbed.yaml')
        self.connect()
        self.generate_topology_file(file)
        with open(site+'.yaml') as f:
            o = yaml.safe_load(f)
        self.l2_switches = o['sites'][site]['devices']['l2_switches']
        self.l3_switches = o['sites'][site]['devices']['l3_hops']
        self.layer2_info = {}
        self.layer3_info = {}
        self.l3_interfaces = {}
        self.endpoints = []

    def connect(self):
        for switch in self.switches:
            try:
                dev = self.testbed.devices[switch]
                dev.connect(init_exec_commands=[], init_config_commands=[],learn_hostname=True)
                self.connected_devices.append(switch)
                self.device_connection_objs[switch] = dev
            except:
                print("Can't Connect to Switch " + switch)
                self.not_connected_devices.append(switch)
        print("SSH to following switches was successfull")

        for switch in self.connected_devices:
            print(switch)
        print("SSH to following switches fail")
        for switch in self.not_connected_devices:
                print(switch)

    def process_input_file(self,file):
            print("processing file")

            # Load the Excel workbook
            workbook = load_workbook(file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2,values_only=True):
             if row[0]!= None:
              self.switches.append(row[0])
            
            print("Connecting to switches:")
            #for switch in self.switches:
            #    print(switch)

            # Validate the structure of the input sheet
            expected_headers = ['hostname', 'ip', 'type','os']
            headers = [cell.value.strip().casefold() if cell.value else '' for cell in sheet[1][:4]]

            if headers != expected_headers:
                print("Invalid structure: The Excel file should have the first three columns with headers 'hostname', 'ip', 'type'.")

            valid_types = {'layer2', 'layer3', 'layer2_layer3'}
            for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
                cell_values = [val.strip().casefold() if val else '' for val in row]
                if not any(cell_value in valid_types for cell_value in cell_values):
                    print("Invalid 'type' value(s) in row ")

            # Create a new Excel workbook for the output
            new_workbook = openpyxl.Workbook()
            new_sheet = new_workbook.active

            # Add new column headers
            new_columns = ['hostname', 'ip', 'username', 'password', 'protocol', 'os']
            new_sheet.append(new_columns)
          
            # Copy data and populate values in the new sheet
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_list = list(row)
                ip = row_list[0]
                hostname = row_list[1]
                os = row_list[3] 
            
                if str(os) == "None":
                    os = "iosxe"
                new_row = []
                new_row.append(ip) 
                new_row.append(hostname) 
                new_row.append('ignore') 
                new_row.append('ignore')
                new_row.append('ssh')
                new_row.append(os)
                
                new_sheet.append(new_row)

            # Save the new Excel file as "inventory.xlsx"
            new_file_name = 'inventory.xlsx'
            new_workbook.save(new_file_name)

            return f"New Excel file '{new_file_name}' created successfully."



    def build_testbed(self):
      
        os.system("pyats create testbed file --path inventory.xlsx --output testbed.yaml")
        with open("testbed.yaml") as r:
            dict = yaml.safe_load(r)
            dict["testbed"] = {'credentials': {'default': {'password': "%ASK{}",'username': "%ASK{}", }}}
        for device, device_data in dict["devices"].items():
            device_data.pop('credentials')
            if ssh_options :
             device_data["connections"]["cli"]["ssh_options"] = ssh_options
      
        with open("testbed.yaml", 'w') as w:
            yaml.dump(dict, w)
        os.system("rm inventory.xlsx")

    def build_offline_testbed(self):
      
        os.system("pyats create testbed file --path inventory.xlsx --output testbed.yaml")
        with open("offline_sites/mock_file.yaml") as f:
            device_mock_file_dict = yaml.safe_load(f)
        with open("testbed.yaml") as r:
            dict = yaml.safe_load(r)
            dict["testbed"] = {'credentials': {'default': {'username': "cisco", 'password': "cisco"}}}
        for device, device_data in dict["devices"].items():
            device_data.pop('credentials')
            device_data["connections"].pop('cli')
            dict["devices"][device]["connections"].update({"default":{"class":"unicon.Unicon"}})
            if device_data["os"] == "iosxe":
             dict["devices"][device]["connections"].update({"a":{"command" : "mock_device_cli --os iosxe --mock_data_dir ./offline_sites/"+site+"/"+device+" --state login","protocol":"unknown"}})
            elif device_data["os"] == "nxos":
             dict["devices"][device]["connections"].update({"a":{"command" : "mock_device_cli --os nxos --mock_data_dir ./offline_sites/"+site+"/"+device+" --state login","protocol":"unknown"}})   
            else:
                print("Unknown Device")
            device_mock_file_dict["exec"]["prompt"] = device+"#"
            device_mock_file_dict["prompt"] = device+"(config)#"
            device_mock_file_dict["config_line"]["prompt"] = device+"(config-line)"
            try:
             with open("./offline_sites/"+site+"/"+device+"/"+device+".yaml","w") as f:
                yaml.dump(device_mock_file_dict,f)
            except:
                print("No outputs found for device "+ device)
                log("No outputs found for device "+ device)
        with open("testbed.yaml", 'w') as f:
            yaml.dump(dict, f) 
        
           
    def generate_topology_file(self, file):

        # Read data from Excel using openpyxl
        excel_file = file
        site_name = self.site
        wb = load_workbook(file)
        ws = wb.active
        yaml_data = {'sites': {site_name: {'devices': {'l2_switches': [], 'l3_hops': []}}}}
        cdp_neighbor_information = {}
        switches = []
        for row in ws.iter_rows(min_row=2, values_only=True):


            if row[0] != None:
             hostname = row[0]
             switches.append(hostname)

             if hostname in self.connected_devices:
              switch = Switch({'hostname':hostname},self.device_connection_objs[hostname])
              cdp_neighbor_information[hostname]=switch.get_switch_cdp_neighbors()
        #pprint.pprint(cdp_neighbor_information)
        for row in ws.iter_rows(min_row=2, values_only=True):


            # Create the device_data dictionary for the switch

            # print(switch,switch_type)
            # Check the type of the switch and place it accordingly in 'l2_switches', 'l3_hops', or both
            trunks_to_ignore_for_mac_learn = []
            if row[0]!= None:
             switch = row[0]
             if switch in self.connected_devices:
              cdp_neighbors = cdp_neighbor_information[switch]
              for neighbor in cdp_neighbors:
                        print(neighbor,switches)
                        for potential_neighbor in switches:
                         if potential_neighbor in neighbor["neighbor"]  :
                            trunks_to_ignore_for_mac_learn.append(neighbor["local_interface"])


             if 'layer2' in row[2] :
                device_data = {'hostname': switch}
                device_data['trunks_to_ignore_for_mac_learn'] = trunks_to_ignore_for_mac_learn
                yaml_data['sites'][site_name]['devices']['l2_switches'].append(device_data)
             if 'layer3' in row[2] :
                device_data = {'hostname': switch}
                yaml_data['sites'][site_name]['devices']['l3_hops'].append(device_data)

        
        yaml_file = self.site+'.yaml'
        with open(yaml_file, 'w') as f:
            yaml.dump(yaml_data, f)

        print("YAML data has been successfully written to the file:", yaml_file)


    def update_layer2_information(self):
        for l2_switch in self.l2_switches:
            if l2_switch['hostname'] in self.connected_devices:
               log("Fetching mac address table and CDP neighbors from " + str(l2_switch['hostname']))
               switch = Switch(l2_switch,self.device_connection_objs[l2_switch['hostname']])
               layer_2_information = switch.get_layer2_information()
               self.layer2_info.update({l2_switch['hostname']:layer_2_information})
               log("Updated Layer 2 Information from "+str(l2_switch['hostname']))

        for layer2_info, layer2_info_data in self.layer2_info.items():
          for endpoint in layer2_info_data:
            #log("updating endpoint "+endpoint.mac+" in sheet")
            self.endpoints.append(endpoint)

    def update_layer3_information(self):



          for l3_switch in self.l3_switches:
            if l3_switch['hostname'] in self.connected_devices:
              log("Fetching layer 3 info from" + l3_switch['hostname'])
              switch = Switch(l3_switch,self.device_connection_objs[l3_switch['hostname']])
              l3_info = switch.get_layer3_information()
              log("Updating layer 3 info from "+l3_switch['hostname'])
              self.layer3_info.update({l3_switch['hostname']:l3_info})

          for endpoint in self.endpoints:


              for layer3_info,layer3_info_data in self.layer3_info.items():
                  for interface, interface_data in layer3_info_data.items():
                          for neighbor, neighbor_data in interface_data['ipv4']['neighbors'].items():
                              if 'Vlan'  in interface:

                               if endpoint.vlan == interface and endpoint.mac == neighbor_data['link_layer_address']:
                                   endpoint.ip = neighbor_data['ip']
                                   log("Found IP of " + str(endpoint.mac) + '=====>' + str(neighbor_data['ip']) + " in " + str(endpoint.vlan))
                              else:
                               if endpoint.mac == neighbor_data['link_layer_address']:
                                  endpoint.ip = neighbor_data['ip']
                                  log("Found IP of "+str(endpoint.mac)+'=====>'+str(neighbor_data['ip']) +" in "+str(endpoint.vlan))

    def get_l3hop_svis(self):

        for l3_switch in self.l3_switches:

          log("Fetching layer 3 info from " + l3_switch['hostname'])
          if l3_switch['hostname'] in self.connected_devices:
            switch = Switch(l3_switch,self.device_connection_objs[l3_switch['hostname']])
            l3_ips = switch.get_l3_ip_mask()

            for l3_ip in l3_ips:
                try:
                    network_address = ipaddress.IPv4Network(l3_ip, strict=False)
                except:
                    log("Cannot find network of " + l3_ip)
                count = 0
                for endpoint in self.endpoints:

                    try:
                        endpoint_ip = ipaddress.IPv4Address(endpoint.ip)
                        if endpoint_ip in network_address:
                            log(endpoint.ip + " part of " + l3_ip)
                            count = count + 1
                    except:
                        pass
                l3_ips[l3_ip]["endpoint_count"] = count

            self.l3_interfaces[l3_switch["hostname"]] = l3_ips
            log("Updated L3 interfaces for " + l3_switch["hostname"])
            log(self.l3_interfaces[l3_switch["hostname"]])


    def generate_report(self):
       self.update_layer2_information()
       self.update_layer3_information()
       self.get_l3hop_svis()

     
       with open(report_folder+"/"+self.site+"/"+self.site+'_endpoints_'+date_time_now+'.csv', 'w') as f:
           log("writing all endpoints to endpoints.csv")
           fieldnames = ['SWITCH','MAC','INTERFACE','INTERFACE_SPEED','INTERFACE_TYPE','VLAN','IP','VENDOR','CDP_PLATFORM','CDP_HOSTNAME','STATIC IP or DHCP','Cisco Comments','Customer Comments','Post-Migration-IP','Critical Endpoint']
           writer = csv.DictWriter(f, fieldnames=fieldnames)
           writer.writeheader()
           for endpoint in self.endpoints :
               dict = {'SWITCH': endpoint.switch,
                       'MAC': endpoint.mac,
                       'INTERFACE': endpoint.interface,
                       'INTERFACE_SPEED': endpoint.interface_speed,
                       'INTERFACE_TYPE': endpoint.interface_type,
                       'VLAN': endpoint.vlan.split('Vlan')[1],
                       'IP': endpoint.ip,
                       'VENDOR': endpoint.vendor,
                       'CDP_PLATFORM': endpoint.cdp_platform,
                       'CDP_HOSTNAME': endpoint.cdp_hostname,
                       'STATIC IP or DHCP': ' ',
                       'Cisco Comments': ' ',
                       'Customer Comments': '',
                       'Post-Migration-IP': ' ',
                       'Critical Endpoint': ' '
                       }
               #log("writing endpoint information about "+endpoint.mac)
               writer.writerow(dict)
       self.csv_to_xlsx(report_folder + '/' + self.site +'/'+self.site+ '_endpoints_'+date_time_now+'.csv',
                        report_folder + '/' + self.site +'/'+self.site+ '_endpoints_'+date_time_now+'.xlsx')
       with open(report_folder + '/' + self.site +'/'+self.site +'_l3_interfaces_'+date_time_now+'.csv', 'w') as f:
           log("writing all l3 interfaces to l3_interfaces.csv")
           fieldnames = ['SWITCH', 'INTERFACE', 'DESCRIPTION', 'IP', "ENDPOINT_COUNT"]
           writer = csv.DictWriter(f, fieldnames=fieldnames)
           writer.writeheader()
           for key, data in self.l3_interfaces.items():

               for key1, data1 in data.items():
                   dict = {'SWITCH': key,
                           'INTERFACE': data1["interface"],
                           'DESCRIPTION': data1["description"],
                           'IP': key1,
                           'ENDPOINT_COUNT': data1["endpoint_count"]
                           }

                   # log("writing endpoint information about "+endpoint.mac)
                   writer.writerow(dict)
       self.csv_to_xlsx(report_folder + '/' + self.site + '/'+self.site+'_l3_interfaces_'+date_time_now+'.csv',
                        report_folder + '/' + self.site + '/'+self.site+'_l3_interfaces_'+date_time_now+'.xlsx')
       os.system("rm "+report_folder + '/' + self.site + '/'+'*.csv')
       endpoint_report = report_folder + '/' + self.site +'/'+self.site+ '_endpoints_'+date_time_now+'.xlsx'
       l3_report = report_folder + '/' + self.site + '/'+self.site+'_l3_interfaces_'+date_time_now+'.xlsx'
       output_file = report_folder + '/' + self.site + '/'+self.site+"_"+date_time_now 
       
       self.merge_excel_sheets(endpoint_report,l3_report,output_file+".xlsx")
       os.system(f"rm {l3_report}")
       os.system(f"rm {endpoint_report}")
       print("Reported generated successfully ")
       print("SSH to following switches was successfull")
       for switch in self.connected_devices:
           print(switch)
       print("SSH to following switches fail")
       for switch in self.not_connected_devices:
           print(switch)

       if args.webex:
        self.publish_gen_configs()




    def csv_to_xlsx(self, csv_file, xlsx_file):
    
    
        try:
            # Create a new Workbook and select the active sheet
            wb = Workbook()
            sheet = wb.active

            # Read the CSV file and write its contents to the XLSX file
            with open(csv_file, 'r', newline='') as csv_file:
                csv_reader = csv.reader(csv_file)
                for row_index, row in enumerate(csv_reader, start=1):
                    for col_index, cell_value in enumerate(row, start=1):
                        sheet.cell(row=row_index, column=col_index, value=cell_value)

            # Save the data to the XLSX file
            wb.save(xlsx_file)

            log(f"Conversion successful. Data written to {xlsx_file}")
        except FileNotFoundError:
            log("Error: CSV file not found.")
        except Exception as e:
            log(f"Error: {e}")
    
    def merge_excel_sheets(self,input_file1, input_file2, output_file):
     try:
        # Load the first workbook and get the first sheet
        workbook1 = openpyxl.load_workbook(input_file1)
        sheet1 = workbook1.active

        # Load the second workbook and get the first sheet
        workbook2 = openpyxl.load_workbook(input_file2)
        sheet2 = workbook2.active

        # Create a new workbook
        new_workbook = openpyxl.Workbook()

        # Create the first worksheet "endpoints" and copy the contents
        new_sheet1 = new_workbook.create_sheet(title="endpoints")
        for row in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row, min_col=1, max_col=sheet1.max_column):
            new_sheet1.append([cell.value for cell in row])

        # Create the second worksheet "l3_interfaces" and copy the contents
        new_sheet2 = new_workbook.create_sheet(title="l3_interfaces")
        for row in sheet2.iter_rows(min_row=1, max_row=sheet2.max_row, min_col=1, max_col=sheet2.max_column):
            new_sheet2.append([cell.value for cell in row])

        # Remove the default sheets created by openpyxl
        new_workbook.remove(new_workbook["Sheet"])
        

        # Save the new workbook
        new_workbook.save(output_file)

        print("Merging completed successfully!")

     except Exception as e:
        print(f"Error: {e}")

# Example usage

    def publish_gen_configs(self):
       #os.system("zip -r "+config_dir+".zip "+ config_dir)
      files_to_zip = [report_folder + '/' + self.site +'/'+self.site+'_'+date_time_now+'.xlsx',
                      'Reports/'+site+'/'+site +'_'+date_time_now + '.log'] 
      output_zip_file = report_folder + '/' + self.site + '/'+self.site+'_'+date_time_now+'.zip'

      
      with zipfile.ZipFile(output_zip_file, 'w') as zipf:
        for file in files_to_zip:
         arcname = os.path.basename(file)
         # Add each file to the zip archive
         try:
          zipf.write(file,arcname = arcname)
         except:
           print("file compression failed, file not published") 
           return     
   
      try:
       with open(webex, 'r') as f:    
        o = yaml.full_load(f)["WEBEX"]
      
       try:
    
         webex_api = WebexTeamsAPI(access_token=o["WEBEX_BOT_TOKEN"],proxies={"https":o["PROXY"]})
        
       except:
         webex_api = WebexTeamsAPI(access_token=o["WEBEX_BOT_TOKEN"])
         print("no webex proxy")
       
       webex_api.messages.create(roomId=o["WEBEX_ROOM"], markdown="Publishing endpoint discovery data for site " + site )
       time.sleep(2)
       webex_api.messages.create(roomId=o["WEBEX_ROOM"], files=[output_zip_file])
      
      except:
          print("Error in publishing report")
    def download_file(self):
        try:

         with open(webex, 'r') as f:    
          o = yaml.full_load(f)["WEBEX"]
          api = WebexTeamsAPI(access_token=o["WEBEX_BOT_TOKEN"])
          messages = api.messages.list(roomId=o["1_1_BOT_ROOM_ID"])
          for message in messages:
            if message.files:
              print(message)
              break
          headers = { 'Authorization': f'Bearer {o["WEBEX_BOT_TOKEN"]}'} 
          file_url =  message.files[0]
          response = requests.get(file_url, headers=headers, stream=True)
          file_name = message.text.split(" ")[1]
          with open(file_name,"wb") as file:
           file.write(response.content)
          print(f"Downloaded file {file_name}")
        except Exception as e :
           print(e)
           print(f"Couldn't download file from webex room")
         



        
class Switch():


    def __init__(self,switch,device_obj):

        self.hostname = switch['hostname']
        self.dev = device_obj
        self.endpoints = []
        self.layer3_info = {}
        self.l3_interfaces = {}
        self.cdp_neighbors = []
        try:
         self.ignore_trunks = switch['trunks_to_ignore_for_mac_learn']
        except:
          pass

    def get_layer2_information(self):
        self.get_interface_status()
        self.get_mac_address_table()
        self.get_cdp_neighbors()

        return self.endpoints

    def get_layer3_information(self):
        #with open(device_folder + '/' + self.hostname + '/show_arp.txt', encoding='utf8',errors='ignore') as f:
        #with open(device_folder + '/' + self.hostname + '/show_arp.txt') as f:
        #with open(device_folder + '/'+self.hostname + '/show_arp.txt', newline='', encoding='utf16') as f:
        #    data = f.read()
        parsed_output = self.dev.parse('show ip arp')
        log(parsed_output)
        return(parsed_output['interfaces'])
        for interface, interface_data in parsed_output['interfaces'].items():
            for neighbor, neighbor_data in interface_data['ipv4']['neighbors'].items():
                  self.layer3_info[neighbor_data['link_layer_address']] = neighbor_data['ip']
        return self.layer3_info

    def get_mac_address_table(self):
        log("Parsing show mac address-table from " + self.hostname)
        pprint.pprint("Parsing show mac address-table from "+ self.hostname)
        parsed_output = self.dev.parse('show mac address-table',)
        pprint.pprint(parsed_output)
        log(parsed_output)
        for vlan , vlan_data in parsed_output['mac_table']['vlans'].items():
            #log("ignoring internal mac addresses shown as CPU ")
            if 'all' not in vlan:
                for mac_address, mac_address_data in vlan_data['mac_addresses'].items():

                    try:
                        for interface in mac_address_data["interfaces"]:
                            if 'thernet' in interface or 'Gig' in interface or 'Port-channel' in interface :

                                log("found physical interface ")
                                physical_interface = interface
                                # log("Ignoring duplicate mac addresses learnt over trunk interfaces ")
                                if physical_interface not in self.ignore_trunks:

                                    endpoint = Endpoint(mac_address)
                                    endpoint.interface = physical_interface
                                    endpoint.vlan = 'Vlan' + vlan
                                    endpoint.switch = self.hostname
                                    vlan_field = self.interfaces['interfaces'][physical_interface]['vlan']
                                    try:
                                       try:
                                        int(vlan_field)
                                        endpoint.interface_type = 'access'
                                       except:
                                        endpoint.interface_type = vlan_field
                                    
                                       endpoint.interface_speed = self.interfaces['interfaces'][physical_interface]['port_speed']
                                    except:
                                        endpoint.interface_speed = "not found"
                                        endpoint.interface_type = "not found"
                                   
                                    try:
                                        endpoint.vendor = MacLookup().lookup(mac_address)
                                    except:
                                        endpoint.vendor = 'Vendor Not Found'
                                    # log("Collecting information about " + mac_address +"__"+ endpoint.vendor+"on " + "Vlan" + vlan)
                                    self.endpoints.append(endpoint)
                    except:
                        log("Did not find any drop for "+ mac_address)

                        endpoint = Endpoint(mac_address)
                        endpoint.interface = "Interface Not Found/ Drop"
                        endpoint.vlan = 'Vlan' + vlan
                        endpoint.switch = self.hostname
                        endpoint.interface_speed = "Interface Not Found/ Drop"
                        endpoint.interface_type = "Interface Not Found/ Drop"
                        try:
                            endpoint.vendor = MacLookup().lookup(mac_address)
                        except:
                            endpoint.vendor = 'Vendor Not Found'
                            # log("Collecting information about " + mac_address +"__"+ endpoint.vendor+"on " + "Vlan" + vlan)
                        self.endpoints.append(endpoint)


    def get_cdp_neighbors(self):
        try:
            log("Parsing show cdp neighbor from " + self.hostname)
            file_location = 'configs/' + self.hostname + '/show_cdp_neighbors.txt'
            #with open(device_folder + '/' + self.hostname + '/show_cdp_neighbors.txt') as f:
                # with open(device_folder + '/'+self.hostname + '/show_cdp_neighbors.txt', newline='', encoding='utf16') as f:
            #    data = f.read()
            log("printing cdp neighbor")
            # log(data)
            parsed_output = self.dev.parse('show cdp neighbor')
            # pprint.pprint("Parsing show cdp neighbor from " + self.hostname)
            log(parsed_output)
            for endpoint in self.endpoints:
                for cdp_neighbor_index, cdp_neighbor_data in parsed_output['cdp']['index'].items():
                    if cdp_neighbor_data['local_interface'] == endpoint.interface:
                        endpoint.cdp_platform = cdp_neighbor_data['platform']
                        endpoint.cdp_hostname = cdp_neighbor_data['device_id']
                        log("Updating cdp information for " + endpoint.mac + "as " + endpoint.cdp_platform)
        except Exception as e:
            log("Found Exception in CDP ")
            log(str(e))


    def get_interface_status(self):
            log("Parsing show interface status from " + self.hostname)
            file_location = 'configs/' + self.hostname + '/show_interface_status.txt'
            log("printing interface")
            # log(data)
            pprint.pprint("Parsing show interface status from " + self.hostname)
            if self.dev.os == 'iosxe':
             self.interfaces = self.dev.parse('show interfaces status')
            if self.dev.os == 'nxos':
             self.interfaces = self.dev.parse('show interface status')
            log(self.interfaces)



    def get_switch_cdp_neighbors(self):
            pprint.pprint("Parsing show cdp neighbor from " + self.hostname)
            parsed_output = self.dev.parse('show cdp neighbor')
            for cdp_neighbor_index, cdp_neighbor_data in parsed_output['cdp']['index'].items():
                    local_interface = cdp_neighbor_data['local_interface']
                    cdp_hostname = cdp_neighbor_data['device_id']
                    self.cdp_neighbors.append({"neighbor":cdp_hostname,"local_interface": local_interface})
            return self.cdp_neighbors

    def get_l3_ip_mask(self):
        log("Parsing show interface from " + self.hostname)

        # log(data)
        pprint.pprint("Parsing show  interface from " + self.hostname)
        if self.dev.os == 'iosxe':
         detail_interfaces = self.dev.parse('show interfaces',timeout = 500 )
        if self.dev.os == 'nxos':
         detail_interfaces = self.dev.parse('show interface', timeout = 500 )   
        log(detail_interfaces)
        for key, data in detail_interfaces.items():

            interface = key
            description = None
            ip_addresses = None
            try:
                ip_addresses = data["ipv4"]
            except:
                pass
            try:
                description = data["description"]
            except:
                pass
            if ip_addresses:
                for ip_address in ip_addresses:
                    self.l3_interfaces[ip_address] = {"interface": interface, "description": description}

        return self.l3_interfaces


class Endpoint():

    def __init__(self,mac):
        self.mac = mac
        self.vlan = None
        self.ip = None
        self.interface = None
        self.switch = None
        self.entry_type = None
        self.cdp_hostname = None
        self.device_type = None
        self.cdp_platform = None
        self.vendor = None
        self.interface_speed = None
        self.interface_type = None

    def info(self):
        endpoint = { 'mac' : self.mac,  'vlan' : self.vlan, 'vendor' : self.vendor}
        return endpoint

def log(message):
        
        with open('Reports/'+site+'/'+site +'_'+date_time_now + '.log', 'a') as f:
            f.write(
                "--------------------------------------------------------------------------------------------------------------------- \n")
            if type(message) != str:
                f.write(pprint.pformat(message))
                f.write("\n")

            else:
                f.write(message)
            f.write(
                "\n---------------------------------------------------------------------------------------------------------------------- \n")

if __name__ == '__main__':


    os.system('mkdir Reports')
    os.system("mkdir Reports/"+site)
    report_folder = 'Reports'
    discover = Discover(args.site,args.file)
    discover.generate_report()

