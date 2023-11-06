#!/usr/bin/env python
__version__ = "0.0.1"
__author__ = "Utkarsh Mahar"
__author_email__ = "umahar@cisco.com"
__copyright__ = "Copyright (c) 2021 Cisco Systems. All rights reserved."
__license__ = "MIT"

import csv
import os
import pprint
import pprint
import sys
import yaml
from datetime import datetime
from genie.conf.base import Device
from genie.testbed import load
from mac_vendor_lookup import MacLookup
from openpyxl import Workbook, load_workbook
from webexteamssdk import WebexTeamsAPI
import argparse
date_time_now =  datetime.now().strftime("%d_%m_%M")
import openpyxl
import ipaddress
import time
parser = argparse.ArgumentParser()
parser.add_argument('-s', '--site', type=str, required=True, help="bucket")
parser.add_argument('-f', '--file', type=str, required=True, help="Job")
parser.add_argument('-w', '--webex', type=str, required=False, help="Webex")
args = parser.parse_args()
site = args.site
file = args.file
webex = args.webex

class Discover():

    def __init__(self,site,file):
        self.site = site
        self.switches = []
        self.device_connection_objs = {}
        self.connected_devices = []
        self.not_connected_devices = []
        self.process_input_file(file)
        self.build_testbed()
        self.testbed = load('testbed.yaml')
        self.connect()
        self.generate_topology_file(file)
        with open(site+'.yaml') as f:
            o = yaml.safe_load(f)
        self.l2_switches = o['sites'][site]['devices']['l2_switches']
        self.l3_switches = o['sites'][site]['devices']['l3_hops']
        self.layer2_info = {}
        self.layer3_info = {}
        self.l3_interfaces = {}
        self.endpoints = []

    def connect(self):
        for switch in self.switches:
            try:
                dev = self.testbed.devices[switch]
                dev.connect(init_exec_commands=[], init_config_commands=[],learn_hostname=True)
                self.connected_devices.append(switch)
                self.device_connection_objs[switch] = dev
            except:
                print("Can't Connect to Switch " + switch)
                self.not_connected_devices.append(switch)
        print("SSH to following switches was successfull")

        for switch in self.connected_devices:
            print(switch)
        print("SSH to following switches fail")
        for switch in self.not_connected_devices:
                print(switch)

    def process_input_file(self,file):
            print("processing file")

            # Load the Excel workbook
            workbook = load_workbook(file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2,values_only=True):
             if row[0]!= None:
              self.switches.append(row[0])

            print("Connecting to switches:")
            for switch in self.switches:
                print(switch)

            # Validate the structure of the input sheet
            expected_headers = ['hostname', 'ip', 'type']
            headers = [cell.value.strip().casefold() if cell.value else '' for cell in sheet[1][:3]]

            if headers != expected_headers:
                return "Invalid structure: The Excel file should have the first three columns with headers 'hostname', 'ip', 'type'."

            valid_types = {'layer2', 'layer3', 'layer2_layer3'}
            for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
                cell_values = [val.strip().casefold() if val else '' for val in row]
                if not any(cell_value in valid_types for cell_value in cell_values):
                    return f"Invalid 'type' value(s) in row "

            # Create a new Excel workbook for the output
            new_workbook = openpyxl.Workbook()
            new_sheet = new_workbook.active

            # Add new column headers
            new_columns = ['hostname', 'ip', 'username', 'password', 'protocol', 'os']
            new_sheet.append(new_columns)

            # Copy data and populate values in the new sheet
            for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
                new_row = list(row) + ['ignore', 'ignore', 'ssh', 'iosxe']
                new_sheet.append(new_row)

            # Save the new Excel file as "inventory.xlsx"
            new_file_name = 'inventory.xlsx'
            new_workbook.save(new_file_name)

            return f"New Excel file '{new_file_name}' created successfully."




    def build_testbed(self):
        os.system("pyats create testbed file --path inventory.xlsx --output testbed.yaml")
        with open("testbed.yaml") as r:
            dict = yaml.safe_load(r)
            dict["testbed"] = {'credentials': {'default': {'username': "%ASK{}", 'password': "%ASK{}"}}}
        for device, device_data in dict["devices"].items():
            device_data.pop('credentials')
        linux_device = {'local_linux': {'os': 'linux', 'type': 'linux', 'connections': {
            'cli': {'protocols': 'ssh', 'ip': '127.0.0.1', 'command': 'bash'}}}}
        dict['devices'].update(linux_device)
        with open("testbed.yaml", 'w') as w:
            yaml.dump(dict, w)

    def generate_topology_file(self, file):

        # Read data from Excel using openpyxl
        excel_file = file
        site_name = self.site
        wb = load_workbook(file)
        ws = wb.active
        yaml_data = {'sites': {site_name: {'devices': {'l2_switches': [], 'l3_hops': []}}}}
        cdp_neighbor_information = {}
        switches = []
        for row in ws.iter_rows(min_row=2, values_only=True):


            if row[0] != None:
             hostname = row[0]
             switches.append(hostname)

             if hostname in self.connected_devices:
              switch = Switch({'hostname':hostname},self.device_connection_objs[hostname])
              cdp_neighbor_information[hostname]=switch.get_switch_cdp_neighbors()
        #pprint.pprint(cdp_neighbor_information)
        for row in ws.iter_rows(min_row=2, values_only=True):


            # Create the device_data dictionary for the switch

            # print(switch,switch_type)
            # Check the type of the switch and place it accordingly in 'l2_switches', 'l3_hops', or both
            trunks_to_ignore_for_mac_learn = []
            if row[0]!= None:
             switch = row[0]
             if switch in self.connected_devices:
              cdp_neighbors = cdp_neighbor_information[switch]
              for neighbor in cdp_neighbors:
                        print(neighbor,switches)
                        for potential_neighbor in switches:
                         if potential_neighbor in neighbor["neighbor"]  :
                            trunks_to_ignore_for_mac_learn.append(neighbor["local_interface"])


             if 'layer2' in row[2] :
                device_data = {'hostname': switch}
                device_data['trunks_to_ignore_for_mac_learn'] = trunks_to_ignore_for_mac_learn
                yaml_data['sites'][site_name]['devices']['l2_switches'].append(device_data)
             if 'layer3' in row[2] :
                device_data = {'hostname': switch}
                yaml_data['sites'][site_name]['devices']['l3_hops'].append(device_data)

        pprint.pprint(yaml_data)
        yaml_file = self.site+'.yaml'
        with open(yaml_file, 'w') as f:
            yaml.dump(yaml_data, f)

        print("YAML data has been successfully written to the file:", yaml_file)


    def update_layer2_information(self):
        for l2_switch in self.l2_switches:
            if l2_switch['hostname'] in self.connected_devices:
               log("Fetching mac address table and CDP neighbors from " + str(l2_switch['hostname']))
               switch = Switch(l2_switch,self.device_connection_objs[l2_switch['hostname']])
               layer_2_information = switch.get_layer2_information()
               self.layer2_info.update({l2_switch['hostname']:layer_2_information})
               log("Updated Layer 2 Information from "+str(l2_switch['hostname']))

        for layer2_info, layer2_info_data in self.layer2_info.items():
          for endpoint in layer2_info_data:
            #log("updating endpoint "+endpoint.mac+" in sheet")
            self.endpoints.append(endpoint)

    def update_layer3_information(self):



          for l3_switch in self.l3_switches:
            if l3_switch['hostname'] in self.connected_devices:
              log("Fetching layer 3 info from" + l3_switch['hostname'])
              switch = Switch(l3_switch,self.device_connection_objs[l3_switch['hostname']])
              l3_info = switch.get_layer3_information()
              log("Updating layer 3 info from "+l3_switch['hostname'])
              self.layer3_info.update({l3_switch['hostname']:l3_info})

          for endpoint in self.endpoints:


              for layer3_info,layer3_info_data in self.layer3_info.items():
                  for interface, interface_data in layer3_info_data.items():
                          for neighbor, neighbor_data in interface_data['ipv4']['neighbors'].items():
                              if 'Vlan'  in interface:

                               if endpoint.vlan == interface and endpoint.mac == neighbor_data['link_layer_address']:
                                   endpoint.ip = neighbor_data['ip']
                                   log("Found IP of " + str(endpoint.mac) + '=====>' + str(neighbor_data['ip']) + " in " + str(endpoint.vlan))
                              else:
                               if endpoint.mac == neighbor_data['link_layer_address']:
                                  endpoint.ip = neighbor_data['ip']
                                  log("Found IP of "+str(endpoint.mac)+'=====>'+str(neighbor_data['ip']) +" in "+str(endpoint.vlan))

    def get_l3hop_svis(self):

        for l3_switch in self.l3_switches:

          log("Fetching layer 3 info from " + l3_switch['hostname'])
          if l3_switch['hostname'] in self.connected_devices:
            switch = Switch(l3_switch,self.device_connection_objs[l3_switch['hostname']])
            l3_ips = switch.get_l3_ip_mask()

            for l3_ip in l3_ips:
                try:
                    network_address = ipaddress.IPv4Network(l3_ip, strict=False)
                except:
                    log("Cannot find network of " + l3_ip)
                count = 0
                for endpoint in self.endpoints:

                    try:
                        endpoint_ip = ipaddress.IPv4Address(endpoint.ip)
                        if endpoint_ip in network_address:
                            log(endpoint.ip + " part of " + l3_ip)
                            count = count + 1
                    except:
                        pass
                l3_ips[l3_ip]["endpoint_count"] = count

            self.l3_interfaces[l3_switch["hostname"]] = l3_ips
            log("Updated L3 interfaces for " + l3_switch["hostname"])
            log(self.l3_interfaces[l3_switch["hostname"]])


    def generate_report(self):
       self.update_layer2_information()
       self.update_layer3_information()
       self.get_l3hop_svis()


       with open(report_folder+'/'+self.site+'_endpoints.csv', 'w') as f:
           log("writing all endpoints to endpoints.csv")
           fieldnames = ['SWITCH','MAC','INTERFACE','INTERFACE_SPEED','INTERFACE_TYPE','VLAN','IP','VENDOR','CDP_PLATFORM','CDP_HOSTNAME','STATIC IP or DHCP','Cisco Comments','Customer Comments','Post-Migration-IP','Critical Endpoint']
           writer = csv.DictWriter(f, fieldnames=fieldnames)
           writer.writeheader()
           for endpoint in self.endpoints :
               dict = {'SWITCH': endpoint.switch,
                       'MAC': endpoint.mac,
                       'INTERFACE': endpoint.interface,
                       'INTERFACE_SPEED': endpoint.interface_speed,
                       'INTERFACE_TYPE': endpoint.interface_type,
                       'VLAN': endpoint.vlan.split('Vlan')[1],
                       'IP': endpoint.ip,
                       'VENDOR': endpoint.vendor,
                       'CDP_PLATFORM': endpoint.cdp_platform,
                       'CDP_HOSTNAME': endpoint.cdp_hostname,
                       'STATIC IP or DHCP': ' ',
                       'Cisco Comments': ' ',
                       'Customer Comments': '',
                       'Post-Migration-IP': ' ',
                       'Critical Endpoint': ' '
                       }
               #log("writing endpoint information about "+endpoint.mac)
               writer.writerow(dict)
       self.csv_to_xlsx(report_folder + '/' + self.site + '_endpoints.csv',
                        report_folder + '/' + self.site + '_endpoints.xlsx')
       with open(report_folder + '/' + self.site + '_l3_interfaces.csv', 'w') as f:
           log("writing all l3 interfaces to l3_interfaces.csv")
           fieldnames = ['SWITCH', 'INTERFACE', 'DESCRIPTION', 'IP', "ENDPOINT_COUNT"]
           writer = csv.DictWriter(f, fieldnames=fieldnames)
           writer.writeheader()
           for key, data in self.l3_interfaces.items():

               for key1, data1 in data.items():
                   dict = {'SWITCH': key,
                           'INTERFACE': data1["interface"],
                           'DESCRIPTION': data1["description"],
                           'IP': key1,
                           'ENDPOINT_COUNT': data1["endpoint_count"]
                           }

                   # log("writing endpoint information about "+endpoint.mac)
                   writer.writerow(dict)
       self.csv_to_xlsx(report_folder + '/' + self.site + '_l3_interfaces.csv',
                        report_folder + '/' + self.site + '_l3_interfaces.xlsx')

       print("Reported generated successfully ")
       print("SSH to following switches was successfull")
       for switch in self.connected_devices:
           print(switch)
       print("SSH to following switches fail")
       for switch in self.not_connected_devices:
           print(switch)

       if args.webex:
        self.publish_gen_configs()




    def csv_to_xlsx(self, csv_file, xlsx_file):
        try:
            # Create a new Workbook and select the active sheet
            wb = Workbook()
            sheet = wb.active

            # Read the CSV file and write its contents to the XLSX file
            with open(csv_file, 'r', newline='') as csv_file:
                csv_reader = csv.reader(csv_file)
                for row_index, row in enumerate(csv_reader, start=1):
                    for col_index, cell_value in enumerate(row, start=1):
                        sheet.cell(row=row_index, column=col_index, value=cell_value)

            # Save the data to the XLSX file
            wb.save(xlsx_file)

            log(f"Conversion successful. Data written to {xlsx_file}")
        except FileNotFoundError:
            log("Error: CSV file not found.")
        except Exception as e:
            log(f"Error: {e}")
    def publish_gen_configs(self):
       #os.system("zip -r "+config_dir+".zip "+ config_dir)
      try:
       with open(webex, 'r') as f:
           o = yaml.full_load(f)["WEBEX"]
       if o["PROXY"]:
        webex_api = WebexTeamsAPI(access_token=o["WEBEX_BOT_TOKEN"],proxies={"https":o["PROXY"]})
       else:
         webex_api = WebexTeamsAPI(access_token=o["WEBEX_BOT_TOKEN"])

       #webex_api.messages.create(roomId=o["WEBEX_ROOM"], markdown="Publishing Test Result ")
       webex_api.messages.create(roomId=o["WEBEX_ROOM"], markdown="Publishing endpoint discovery data for site " + site )
       time.sleep(2)
       webex_api.messages.create(roomId=o["WEBEX_ROOM"], files=['Reports/'+site+'_endpoints.xlsx'])
       time.sleep(3)
       webex_api.messages.create(roomId=o["WEBEX_ROOM"], files=['Reports/'+site+'_l3_interfaces.xlsx'])
       time.sleep(3)
       webex_api.messages.create(roomId=o["WEBEX_ROOM"], files=['Reports/'+site+'_endpoints_' + date_time_now + '.log'])

      except:
          print("Error in publishing report")

class Switch():


    def __init__(self,switch,device_obj):

        self.hostname = switch['hostname']
        self.dev = device_obj
        self.endpoints = []
        self.layer3_info = {}
        self.l3_interfaces = {}
        self.cdp_neighbors = []
        try:
         self.ignore_trunks = switch['trunks_to_ignore_for_mac_learn']
        except:
          pass

    def get_layer2_information(self):
        self.get_interface_status()
        self.get_mac_address_table()
        self.get_cdp_neighbors()

        return self.endpoints

    def get_layer3_information(self):
        #with open(device_folder + '/' + self.hostname + '/show_arp.txt', encoding='utf8',errors='ignore') as f:
        #with open(device_folder + '/' + self.hostname + '/show_arp.txt') as f:
        #with open(device_folder + '/'+self.hostname + '/show_arp.txt', newline='', encoding='utf16') as f:
        #    data = f.read()
        parsed_output = self.dev.parse('show arp')
        log(parsed_output)
        return(parsed_output['interfaces'])
        for interface, interface_data in parsed_output['interfaces'].items():
            for neighbor, neighbor_data in interface_data['ipv4']['neighbors'].items():
                  self.layer3_info[neighbor_data['link_layer_address']] = neighbor_data['ip']
        return self.layer3_info

    def get_mac_address_table(self):
        log("Parsing show mac address-table from " + self.hostname)
        #with open(device_folder + '/' + self.hostname + '/show_mac_address-table.txt', 'rb') as f:

        #with open(device_folder+'/'+self.hostname+'/show_mac_address-table.txt') as f:
        #with open(device_folder + '/' + self.hostname + '/show_mac_address-table.txt', encoding='utf8', errors='ignore') as f:
        #with open(device_folder + '/'+self.hostname + '/show_mac_address-table.txt', newline='', encoding='utf16') as f:
        #    data = f.read()
        pprint.pprint("Parsing show mac address-table from "+ self.hostname)
        parsed_output = self.dev.parse('show mac address-table',)
        pprint.pprint(parsed_output)
        log(parsed_output)
        for vlan , vlan_data in parsed_output['mac_table']['vlans'].items():
            #log("ignoring internal mac addresses shown as CPU ")
            if 'all' not in vlan:
                for mac_address, mac_address_data in vlan_data['mac_addresses'].items():

                    try:
                        for interface in mac_address_data["interfaces"]:
                            if 'thernet' in interface or 'Gig' in interface:

                                log("found physical interface ")
                                physical_interface = interface
                                # log("Ignoring duplicate mac addresses learnt over trunk interfaces ")
                                if physical_interface not in self.ignore_trunks:

                                    endpoint = Endpoint(mac_address)
                                    endpoint.interface = physical_interface
                                    endpoint.vlan = 'Vlan' + vlan
                                    endpoint.switch = self.hostname
                                    try:
                                        if 'trunk' != self.interfaces['interfaces'][physical_interface]['vlan'] or 'routed' != self.interfaces['interfaces'][physical_interface]['vlan']:
                                            endpoint.interface_type = 'access'
                                        else:
                                            endpoint.interface_type = self.interfaces['interfaces'][physical_interface]['vlan']
                                        endpoint.interface_speed = self.interfaces['interfaces'][physical_interface]['port_speed']
                                    except:
                                        endpoint.interface_speed = "not found"
                                        endpoint.interface_type = "not found"
                                    try:
                                        endpoint.vendor = MacLookup().lookup(mac_address)
                                    except:
                                        endpoint.vendor = 'Vendor Not Found'
                                    # log("Collecting information about " + mac_address +"__"+ endpoint.vendor+"on " + "Vlan" + vlan)
                                    self.endpoints.append(endpoint)
                    except:
                        log("Did not find any drop for "+ mac_address)

                        endpoint = Endpoint(mac_address)
                        endpoint.interface = "Interface Not Found/ Drop"
                        endpoint.vlan = 'Vlan' + vlan
                        endpoint.switch = self.hostname
                        endpoint.interface_speed = "Interface Not Found/ Drop"
                        endpoint.interface_type = "Interface Not Found/ Drop"
                        try:
                            endpoint.vendor = MacLookup().lookup(mac_address)
                        except:
                            endpoint.vendor = 'Vendor Not Found'
                            # log("Collecting information about " + mac_address +"__"+ endpoint.vendor+"on " + "Vlan" + vlan)
                        self.endpoints.append(endpoint)


    def get_cdp_neighbors(self):
        try:
            log("Parsing show cdp neighbor from " + self.hostname)
            file_location = 'configs/' + self.hostname + '/show_cdp_neighbors.txt'
            #with open(device_folder + '/' + self.hostname + '/show_cdp_neighbors.txt') as f:
                # with open(device_folder + '/'+self.hostname + '/show_cdp_neighbors.txt', newline='', encoding='utf16') as f:
            #    data = f.read()
            log("printing cdp neighbor")
            # log(data)
            parsed_output = self.dev.parse('show cdp neighbor')
            # pprint.pprint("Parsing show cdp neighbor from " + self.hostname)
            log(parsed_output)
            for endpoint in self.endpoints:
                for cdp_neighbor_index, cdp_neighbor_data in parsed_output['cdp']['index'].items():
                    if cdp_neighbor_data['local_interface'] == endpoint.interface:
                        endpoint.cdp_platform = cdp_neighbor_data['platform']
                        endpoint.cdp_hostname = cdp_neighbor_data['device_id']
                        log("Updating cdp information for " + endpoint.mac + "as " + endpoint.cdp_platform)
        except Exception as e:
            log("Found Exception in CDP ")
            log(str(e))


    def get_interface_status(self):
            log("Parsing show interface status from " + self.hostname)
            file_location = 'configs/' + self.hostname + '/show_interface_status.txt'
            log("printing interface")
            # log(data)
            pprint.pprint("Parsing show interface status from " + self.hostname)
            self.interfaces = self.dev.parse('show interfaces status')
            log(self.interfaces)



    def get_switch_cdp_neighbors(self):
            pprint.pprint("Parsing show cdp neighbor from " + self.hostname)
            parsed_output = self.dev.parse('show cdp neighbor')
            for cdp_neighbor_index, cdp_neighbor_data in parsed_output['cdp']['index'].items():
                    local_interface = cdp_neighbor_data['local_interface']
                    cdp_hostname = cdp_neighbor_data['device_id']
                    self.cdp_neighbors.append({"neighbor":cdp_hostname,"local_interface": local_interface})
            return self.cdp_neighbors

    def get_l3_ip_mask(self):
        log("Parsing show interface from " + self.hostname)

        # log(data)
        pprint.pprint("Parsing show  interface from " + self.hostname)
        detail_interfaces = self.dev.parse('show interfaces', )
        log(detail_interfaces)
        for key, data in detail_interfaces.items():

            interface = key
            description = None
            ip_addresses = None
            try:
                ip_addresses = data["ipv4"]
            except:
                pass
            try:
                description = data["description"]
            except:
                pass
            if ip_addresses:
                for ip_address in ip_addresses:
                    self.l3_interfaces[ip_address] = {"interface": interface, "description": description}

        return self.l3_interfaces


class Endpoint():

    def __init__(self,mac):
        self.mac = mac
        self.vlan = None
        self.ip = None
        self.interface = None
        self.switch = None
        self.entry_type = None
        self.cdp_hostname = None
        self.device_type = None
        self.cdp_platform = None
        self.vendor = None
        self.interface_speed = None
        self.interface_type = None

    def info(self):
        endpoint = { 'mac' : self.mac,  'vlan' : self.vlan, 'vendor' : self.vendor}
        return endpoint

def log(message):
        with open('Reports/'+site+'_endpoints_' + date_time_now + '.log', 'a') as f:
            f.write(
                "--------------------------------------------------------------------------------------------------------------------- \n")
            if type(message) != str:
                f.write(pprint.pformat(message))
                f.write("\n")

            else:
                f.write(message)
            f.write(
                "\n---------------------------------------------------------------------------------------------------------------------- \n")

if __name__ == '__main__':


    os.system('mkdir Reports')
    report_folder = 'Reports'
    discover = Discover(args.site,args.file)
    discover.generate_report()
